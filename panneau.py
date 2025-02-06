from io import BytesIO
import streamlit as st
from openpyxl import load_workbook, Workbook

def exporter_donnees_pulsar(wb_source):
    """
    Exporte les données du produit PULSAR ainsi que les options/accessoires
    depuis le fichier Excel chargé par l'utilisateur.
    """

    # Vérifier l'existence de l'onglet "PULSAR"
    if "PULSAR" not in wb_source.sheetnames:
        st.error("Erreur : L'onglet 'PULSAR' n'existe pas dans le fichier.")
        return None

    ws_source = wb_source["PULSAR"]

    # Création du fichier destination
    wb_destination = Workbook()
    ws_destination = wb_destination.active
    ws_destination.title = "Données Copiées"

    # Initialisation d'un dictionnaire pour suivre les titres exportés
    titres_exportes = {}

    # Parcourir la plage A15:A41
    dernier_titre = ""
    ligne_destination = 1

    for row in range(15, 42):  # Lignes 15 à 41
        cell_value = ws_source[f"A{row}"].value  # Colonne A
        if cell_value:
            dernier_titre = cell_value
            if dernier_titre not in titres_exportes:
                titres_exportes[dernier_titre] = True
                ws_destination[f"B{ligne_destination}"] = dernier_titre
                ligne_destination += 1

        if dernier_titre:
            quantite = ws_source[f"B{row}"].value  # Colonne B
            if quantite:
                type_valeur = ws_source[f"C{row}"].value  # Colonne C
                position = ws_source[f"D{row}"].value  # Colonne D
                reference_valeur = ws_source[f"O{row}"].value  # Colonne O

                # Insérer les données dans la feuille destination
                ws_destination[f"A{ligne_destination}"] = str(reference_valeur)  # Stocké en texte
                if type_valeur and position:
                    ws_destination[f"B{ligne_destination}"] = f"{type_valeur} {position}"
                ws_destination[f"C{ligne_destination}"] = quantite
                ligne_destination += 1

    # Cas où il n'y a pas de titre en colonne A
    if not dernier_titre:
        for row in range(15, 42):
            reference_valeur = ws_source[f"O{row}"].value  # Colonne O
            if reference_valeur:
                quantite = ws_source[f"B{row}"].value  # Colonne B
                if quantite:
                    type_valeur = ws_source[f"C{row}"].value  # Colonne C
                    position = ws_source[f"D{row}"].value  # Colonne D

                    # Insérer les données
                    ws_destination[f"A{ligne_destination}"] = str(reference_valeur)
                    if type_valeur and position:
                        ws_destination[f"B{ligne_destination}"] = f"{type_valeur} {position}"
                    ws_destination[f"C{ligne_destination}"] = quantite
                    ligne_destination += 1

    # Traitement des données de la section "Options / Accessoires"
    ligne_depart = ligne_destination + 2
    ws_destination[f"B{ligne_depart - 1}"] = "Accessoires PULSAR"
    ws_destination[f"I{ligne_depart - 1}"] = "T"

    for row in range(47, 70):  # Lignes 47 à 69
        code = ws_source[f"A{row}"].value  # Colonne A
        if code:
            quantite = ws_source[f"O{row}"].value  # Colonne O
            ws_destination[f"A{ligne_depart}"] = code
            ws_destination[f"C{ligne_depart}"] = quantite
            ligne_depart += 1

    return wb_destination


# Interface Streamlit
st.title("Exportation des données PULSAR")

uploaded_file = st.file_uploader("Choisissez un fichier Excel", type=["xlsx", "xlsm"])

if uploaded_file:
    # Charger le fichier Excel dans OpenPyXL
    wb_source = load_workbook(uploaded_file, data_only=True)

    # Traiter les données
    wb_result = exporter_donnees_pulsar(wb_source)

    if wb_result:
        # Sauvegarder en mémoire
        output = BytesIO()
        wb_result.save(output)
        output.seek(0)

        # Bouton de téléchargement
        st.download_button(
            label="Télécharger le fichier transformé",
            data=output,
            file_name="donnees_pulsar.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        st.success("Traitement terminé ! Vous pouvez télécharger le fichier.")
